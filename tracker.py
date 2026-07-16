# tracker.py
from __future__ import annotations
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Iterator
import openpyxl
from openpyxl.cell.cell import Cell

COL_PRIORITY = 1
COL_STATUS = 2
COL_DATE = 3
COL_NOTES = 4
COL_COMPANY = 5
COL_ROLE = 6
COL_CITY = 7
COL_REGION = 8
COL_LINK = 11
COL_RESUME_PATH = 13


@dataclass
class ApplicationRow:
    excel_row: int
    company: str
    title: str
    city: str
    url: str
    priority: str
    status: str
    resume_path: str | None


def iter_applications(
    tracker_path: Path,
    priority_prefix: str = "A",
    status_filter: str = "To apply",
) -> Iterator[ApplicationRow]:
    wb = openpyxl.load_workbook(tracker_path)
    ws = wb["Apply Tracker"]
    for row in range(2, ws.max_row + 1):
        priority = str(ws.cell(row, COL_PRIORITY).value or "")
        status = str(ws.cell(row, COL_STATUS).value or "")
        if not priority.startswith(priority_prefix):
            continue
        if status != status_filter:
            continue
        link_cell = ws.cell(row, COL_LINK)
        url = (link_cell.hyperlink.target or "") if link_cell.hyperlink else ""
        resume_cell = ws.cell(row, COL_RESUME_PATH)
        resume_path = str(resume_cell.value) if resume_cell.value else None
        yield ApplicationRow(
            excel_row=row,
            company=str(ws.cell(row, COL_COMPANY).value or ""),
            title=str(ws.cell(row, COL_ROLE).value or ""),
            city=str(ws.cell(row, COL_CITY).value or ""),
            url=url,
            priority=priority,
            status=status,
            resume_path=resume_path,
        )


def update_status(
    tracker_path: Path,
    excel_row: int,
    status: str,
    notes: str = "",
) -> None:
    wb = openpyxl.load_workbook(tracker_path)
    ws = wb["Apply Tracker"]
    status_cell = ws.cell(excel_row, COL_STATUS)
    assert isinstance(status_cell, Cell)
    status_cell.value = status
    if status == "Applied":
        date_cell = ws.cell(excel_row, COL_DATE)
        assert isinstance(date_cell, Cell)
        date_cell.value = date.today().isoformat()
    if notes:
        notes_cell = ws.cell(excel_row, COL_NOTES)
        assert isinstance(notes_cell, Cell)
        existing = str(notes_cell.value or "")
        notes_cell.value = (existing + " | " + notes).strip(" | ")
    wb.save(tracker_path)


def append_application(
    tracker_path: Path,
    *,
    company: str,
    role: str,
    city: str,
    url: str,
    priority: str = "A — strong fit",
    why_fits: str = "",
) -> int:
    """Append a new 'To apply' row to the Apply Tracker sheet. Returns the new
    Excel row number. Sets the Link cell's hyperlink so iter_applications picks
    up the URL exactly as apply.py expects."""
    wb = openpyxl.load_workbook(tracker_path)
    ws = wb["Apply Tracker"]
    r = ws.max_row + 1
    ws.cell(r, COL_PRIORITY).value = priority
    ws.cell(r, COL_STATUS).value = "To apply"
    ws.cell(r, COL_COMPANY).value = company
    ws.cell(r, COL_ROLE).value = role
    ws.cell(r, COL_CITY).value = city
    if why_fits:
        ws.cell(r, 9).value = why_fits  # "Why it fits" column
    link_cell = ws.cell(r, COL_LINK)
    link_cell.value = "apply ▸"
    link_cell.hyperlink = url
    wb.save(tracker_path)
    return r
