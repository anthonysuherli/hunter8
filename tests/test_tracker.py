# tests/test_tracker.py
import pytest
import openpyxl
from tracker import iter_applications, update_status


@pytest.fixture
def tmp_tracker(tmp_path):
    wb = openpyxl.Workbook()
    wb.create_sheet("Profile").append(["key", "value"])
    ws = wb.create_sheet("Apply Tracker")
    default = wb["Sheet"]
    wb.remove(default)
    headers = ["Priority","Status","Date applied","My notes","Company","Role",
               "City","Region","Why it fits","Flags","Link","Verification","Tailored Resume Path"]
    ws.append(headers)
    ws.append(["A — strong fit","To apply","","","Anthropic","Research Engineer",
               "NYC","NYC","fits","","","","~/career/.../resume-tailored.md"])
    ws.cell(2, 11).hyperlink = "https://job-boards.greenhouse.io/anthropic/jobs/4017331008"
    ws.cell(2, 11).value = "apply ▸"
    ws.append(["B — good fit","To apply","","","Datadog","Staff Scientist",
               "NYC","NYC","fits","","","",""])
    path = tmp_path / "tracker.xlsx"
    wb.save(path)
    return path


def test_iter_applications_returns_a_fit_only(tmp_tracker):
    rows = list(iter_applications(tmp_tracker))
    assert len(rows) == 1
    assert rows[0].company == "Anthropic"


def test_iter_applications_row_has_url(tmp_tracker):
    rows = list(iter_applications(tmp_tracker))
    assert rows[0].url == "https://job-boards.greenhouse.io/anthropic/jobs/4017331008"


def test_update_status_writes_applied(tmp_tracker):
    update_status(tmp_tracker, excel_row=2, status="Applied", notes="auto-submit · greenhouse")
    wb = openpyxl.load_workbook(tmp_tracker)
    ws = wb["Apply Tracker"]
    assert ws.cell(2, 2).value == "Applied"
    assert ws.cell(2, 4).value == "auto-submit · greenhouse"


def test_update_status_writes_date_on_applied(tmp_tracker):
    from datetime import date
    update_status(tmp_tracker, excel_row=2, status="Applied")
    wb = openpyxl.load_workbook(tmp_tracker)
    ws = wb["Apply Tracker"]
    assert ws.cell(2, 3).value == date.today().isoformat()


def test_append_application_adds_row_with_hyperlink(tmp_tracker):
    from tracker import append_application
    append_application(
        tmp_tracker, company="Hebbia", role="Research Engineer", city="SF",
        url="https://job-boards.greenhouse.io/hebbia/jobs/1", priority="A — strong fit",
        why_fits="agentic AI × finance",
    )
    wb = openpyxl.load_workbook(tmp_tracker)
    ws = wb["Apply Tracker"]
    last = ws.max_row
    assert ws.cell(last, 5).value == "Hebbia"      # Company
    assert ws.cell(last, 6).value == "Research Engineer"  # Role
    assert ws.cell(last, 2).value == "To apply"    # Status
    assert ws.cell(last, 1).value == "A — strong fit"     # Priority
    assert ws.cell(last, 11).hyperlink.target == \
        "https://job-boards.greenhouse.io/hebbia/jobs/1"


def test_append_then_iter_returns_it(tmp_tracker):
    from tracker import append_application, iter_applications
    append_application(
        tmp_tracker, company="Rogo", role="ML Engineer", city="NYC",
        url="https://jobs.ashbyhq.com/rogo/1", priority="A — strong fit",
    )
    companies = [r.company for r in iter_applications(tmp_tracker)]
    assert "Rogo" in companies
