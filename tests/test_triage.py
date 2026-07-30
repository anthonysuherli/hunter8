# tests/test_triage.py
import openpyxl

import db as dbmod
import triage
from db import Job


def _tracker(tmp_path):
    wb = openpyxl.Workbook()
    wb.create_sheet("Profile").append(["key", "value"])
    ws = wb.create_sheet("Apply Tracker")
    wb.remove(wb["Sheet"])
    headers = ["Priority", "Status", "Date applied", "My notes", "Company", "Role",
               "City", "Region", "Why it fits", "Flags", "Link", "Verification",
               "Tailored Resume Path"]
    ws.append(headers)
    p = tmp_path / "tracker.xlsx"
    wb.save(p)
    return p


def _scored_job(conn, title="ML Engineer", grade="A"):
    if not dbmod.insert_job(conn, Job(url="https://job-boards.greenhouse.io/acme/jobs/1",
                               company="Acme", title=title, location="Remote US",
                               source="ats:greenhouse", ats="greenhouse", raw_text="d")):
        return
    j = dbmod.jobs_by_status(conn, "discovered")[0]
    dbmod.set_score(conn, j.id, status="scored", grade=grade, reasoning="fit",
                    archetype="lab", comp_signal="$180k", red_flags="[]")


def test_apply_decision_approve_writes_tracker_row(tmp_path):
    tracker = _tracker(tmp_path)
    conn = dbmod.connect(tmp_path / "h.db")
    dbmod.init_db(conn)
    _scored_job(conn)
    job = dbmod.jobs_by_status(conn, "scored")[0]

    triage.apply_decision(conn, job, "a", tracker)

    assert len(dbmod.jobs_by_status(conn, "approved")) == 1
    wb = openpyxl.load_workbook(tracker)
    ws = wb["Apply Tracker"]
    assert ws.cell(ws.max_row, 5).value == "Acme"


def test_apply_decision_skip_and_snooze(tmp_path):
    tracker = _tracker(tmp_path)
    conn = dbmod.connect(tmp_path / "h.db")
    dbmod.init_db(conn)
    _scored_job(conn)
    job = dbmod.jobs_by_status(conn, "scored")[0]
    triage.apply_decision(conn, job, "s", tracker)
    assert len(dbmod.jobs_by_status(conn, "skipped")) == 1

    _scored_job(conn)  # different call reuses same URL → deduped; insert a new one
    dbmod.insert_job(conn, Job(url="https://x/2", company="B", title="AI Engineer",
                               location="Remote US", source="ats:greenhouse",
                               ats="greenhouse", raw_text="d"))
    j2 = dbmod.jobs_by_status(conn, "discovered")[0]
    dbmod.set_score(conn, j2.id, status="scored", grade="B", reasoning="",
                    archetype="", comp_signal="", red_flags="[]")
    triage.apply_decision(conn, dbmod.jobs_by_status(conn, "scored")[0], "z", tracker)
    assert len(dbmod.jobs_by_status(conn, "snoozed")) == 1


def test_priority_from_grade():
    assert triage.priority_from_grade("A").startswith("A")
    assert triage.priority_from_grade("B").startswith("B")
    assert triage.priority_from_grade("C").startswith("C")


def test_approve_ids_writes_rows_and_reports_each(tmp_path):
    tracker = _tracker(tmp_path)
    conn = dbmod.connect(tmp_path / "h.db")
    dbmod.init_db(conn)
    _scored_job(conn)
    job = dbmod.jobs_by_status(conn, "scored")[0]

    results = triage.approve_ids(conn, [job.id], tracker)

    assert results == [{"id": job.id, "ok": True,
                        "detail": "Acme — ML Engineer"}]
    assert len(dbmod.jobs_by_status(conn, "approved")) == 1


def test_approve_ids_reports_an_unknown_id_without_aborting_the_rest(tmp_path):
    tracker = _tracker(tmp_path)
    conn = dbmod.connect(tmp_path / "h.db")
    dbmod.init_db(conn)
    _scored_job(conn)
    job = dbmod.jobs_by_status(conn, "scored")[0]

    results = triage.approve_ids(conn, [99999, job.id], tracker)

    assert results[0]["ok"] is False and "no job" in results[0]["detail"]
    assert results[1]["ok"] is True
    assert len(dbmod.jobs_by_status(conn, "approved")) == 1


def test_approving_an_already_approved_job_is_an_idempotent_no_op(tmp_path):
    tracker = _tracker(tmp_path)
    conn = dbmod.connect(tmp_path / "h.db")
    dbmod.init_db(conn)
    _scored_job(conn)
    job = dbmod.jobs_by_status(conn, "scored")[0]
    triage.approve_ids(conn, [job.id], tracker)

    results = triage.approve_ids(conn, [job.id], tracker)

    assert results[0]["ok"] is False
    assert "already approved" in results[0]["detail"]
    wb = openpyxl.load_workbook(tracker)
    assert wb["Apply Tracker"].max_row == 2   # header + one row, not two


def test_approve_ids_on_an_empty_list_does_nothing(tmp_path):
    tracker = _tracker(tmp_path)
    conn = dbmod.connect(tmp_path / "h.db")
    dbmod.init_db(conn)
    assert triage.approve_ids(conn, [], tracker) == []


def test_a_failed_tracker_write_is_reported_and_does_not_stop_the_rest(monkeypatch, tmp_path):
    """The tracker is an .xlsx the user may have open in Excel, and openpyxl
    cannot write a workbook Excel is holding. That failure has to be visible per
    id — and must not swallow the ids that follow it."""
    def _fail_once(*args, **kwargs):
        raise PermissionError("[Errno 13] Permission denied: 'ML-AI-Roles-Tracker.xlsx'")

    monkeypatch.setattr(triage, "append_application", _fail_once)

    tracker = _tracker(tmp_path)
    conn = dbmod.connect(tmp_path / "h.db")
    dbmod.init_db(conn)
    _scored_job(conn, title="ML Engineer", grade="A")
    # Insert second job with different URL to avoid deduplication
    dbmod.insert_job(conn, Job(url="https://job-boards.greenhouse.io/acme/jobs/2",
                               company="Acme", title="AI Engineer", location="Remote US",
                               source="ats:greenhouse", ats="greenhouse", raw_text="d"))
    j2 = dbmod.jobs_by_status(conn, "discovered")[0]
    dbmod.set_score(conn, j2.id, status="scored", grade="B", reasoning="fit",
                    archetype="lab", comp_signal="$180k", red_flags="[]")

    jobs = dbmod.jobs_by_status(conn, "scored")
    first_id = jobs[0].id
    second_id = jobs[1].id

    results = triage.approve_ids(conn, [first_id, second_id], tracker)

    # Both results should report failure
    assert len(results) == 2
    assert results[0]["id"] == first_id
    assert results[0]["ok"] is False
    assert "tracker write failed" in results[0]["detail"]

    assert results[1]["id"] == second_id
    assert results[1]["ok"] is False
    assert "tracker write failed" in results[1]["detail"]

    # No jobs should have moved to approved
    assert len(dbmod.jobs_by_status(conn, "approved")) == 0


def test_one_failing_id_does_not_block_a_later_good_one(monkeypatch, tmp_path):
    """Verifies the per-id isolation is real, not just that the loop finishes."""
    import tracker as tracker_module
    call_count = [0]  # Use list to allow modification in nested function

    def _fail_first(*args, **kwargs):
        call_count[0] += 1
        if call_count[0] == 1:
            raise PermissionError("[Errno 13] Permission denied: 'ML-AI-Roles-Tracker.xlsx'")
        return tracker_module.append_application(*args, **kwargs)

    monkeypatch.setattr(triage, "append_application", _fail_first)

    tracker = _tracker(tmp_path)
    conn = dbmod.connect(tmp_path / "h.db")
    dbmod.init_db(conn)
    _scored_job(conn, title="ML Engineer", grade="A")
    # Insert second job with different URL to avoid deduplication
    dbmod.insert_job(conn, Job(url="https://job-boards.greenhouse.io/acme/jobs/2",
                               company="Acme", title="AI Engineer", location="Remote US",
                               source="ats:greenhouse", ats="greenhouse", raw_text="d"))
    j2 = dbmod.jobs_by_status(conn, "discovered")[0]
    dbmod.set_score(conn, j2.id, status="scored", grade="B", reasoning="fit",
                    archetype="lab", comp_signal="$180k", red_flags="[]")

    jobs = dbmod.jobs_by_status(conn, "scored")
    first_id = jobs[0].id
    second_id = jobs[1].id

    results = triage.approve_ids(conn, [first_id, second_id], tracker)

    # First should fail, second should succeed
    assert len(results) == 2
    assert results[0]["id"] == first_id
    assert results[0]["ok"] is False
    assert "tracker write failed" in results[0]["detail"]

    assert results[1]["id"] == second_id
    assert results[1]["ok"] is True
    assert "AI Engineer" in results[1]["detail"]

    # Exactly one job should be approved
    approved = dbmod.jobs_by_status(conn, "approved")
    assert len(approved) == 1
    assert approved[0].id == second_id

    # Exactly one new row in tracker (header + 1)
    wb = openpyxl.load_workbook(tracker)
    ws = wb["Apply Tracker"]
    assert ws.max_row == 2
